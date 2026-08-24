from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from pathlib import Path
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

import requests
import shutil
import re
import json
from datetime import datetime
from typing import Optional, List, Dict, Any

# Supabase client removed, using sqlite


# =========================================================
# PATHS
# =========================================================

BASE_DIR = Path(__file__).resolve().parent.parent

DOCUMENTS_DIR = BASE_DIR / "documents"
WEBSITE_DIR = BASE_DIR / "app"

DOCUMENTS_DIR.mkdir(exist_ok=True)


# =========================================================
# SETTINGS
# =========================================================

LLAMA_URL = "http://localhost:8080/v1/chat/completions"

# We will retrieve only a small amount of information
# because your laptop currently uses a 2048 context.
TOP_CHUNKS = 5

MAX_CHUNK_CHARS = 1800


# =========================================================
# FASTAPI
# =========================================================

app = FastAPI(title="My Local AI API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:8000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# DOCUMENT STORAGE
# =========================================================

documents = []
vectorizer = None
document_matrix = None


# =========================================================
# TEXT EXTRACTION
# =========================================================

def extract_pdf(path: Path):
    text = ""

    reader = PdfReader(str(path))

    for page in reader.pages:
        page_text = page.extract_text()

        if page_text:
            text += page_text + "\n"

    return text


def extract_docx(path: Path):
    doc = Document(str(path))

    text = []

    for paragraph in doc.paragraphs:
        if paragraph.text.strip():
            text.append(paragraph.text)

    return "\n".join(text)


def extract_txt(path: Path):
    return path.read_text(
        encoding="utf-8",
        errors="ignore"
    )


def extract_xlsx(path: Path):
    workbook = load_workbook(
        filename=str(path),
        read_only=True,
        data_only=True
    )

    text = []

    for sheet in workbook.worksheets:

        text.append(f"Sheet: {sheet.title}")

        for row in sheet.iter_rows(values_only=True):

            values = []

            for cell in row:

                if cell is not None:
                    values.append(str(cell))

            if values:
                text.append(" | ".join(values))

    return "\n".join(text)


def extract_text(path: Path):

    extension = path.suffix.lower()

    if extension == ".pdf":
        return extract_pdf(path)

    if extension == ".docx":
        return extract_docx(path)

    if extension in [".txt", ".md"]:
        return extract_txt(path)

    if extension == ".xlsx":
        return extract_xlsx(path)

    raise ValueError(
        f"Unsupported file type: {extension}"
    )


# =========================================================
# TEXT CLEANING
# =========================================================

def clean_text(text):

    text = text.replace("\x00", " ")

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip()


# =========================================================
# CHUNKING
# =========================================================

def create_chunks(text):

    text = clean_text(text)

    if not text:
        return []

    chunks = []

    start = 0

    while start < len(text):

        end = start + MAX_CHUNK_CHARS

        chunk = text[start:end]

        if chunk.strip():
            chunks.append(chunk.strip())

        start = end

    return chunks


# =========================================================
# BUILD DOCUMENT INDEX
# =========================================================

def rebuild_index():

    global documents
    global vectorizer
    global document_matrix

    documents = []

    print("\nScanning documents...")

    for path in DOCUMENTS_DIR.rglob("*"):

        if not path.is_file():
            continue

        if path.suffix.lower() not in [
            ".pdf",
            ".docx",
            ".txt",
            ".md",
            ".xlsx"
        ]:
            continue

        try:

            text = extract_text(path)

            chunks = create_chunks(text)

            for number, chunk in enumerate(chunks):

                documents.append({
                    "file": path.name,
                    "path": str(path),
                    "chunk": number,
                    "text": chunk,
                    "modified": path.stat().st_mtime
                })

            print(
                f"Indexed: {path.name} "
                f"({len(chunks)} chunks)"
            )

        except Exception as error:

            print(
                f"Could not read {path.name}: {error}"
            )

    if not documents:

        vectorizer = None
        document_matrix = None

        print("No documents found.")

        return

    texts = [
        item["text"]
        for item in documents
    ]

    vectorizer = TfidfVectorizer(
        stop_words="english",
        ngram_range=(1, 2)
    )

    document_matrix = vectorizer.fit_transform(texts)

    print(
        f"Total indexed chunks: {len(documents)}"
    )


# =========================================================
# SEARCH DOCUMENTS
# =========================================================

def search_documents(question):

    if not documents:
        return []

    query_vector = vectorizer.transform(
        [question]
    )

    scores = cosine_similarity(
        query_vector,
        document_matrix
    )[0]

    ranked = sorted(
        enumerate(scores),
        key=lambda x: x[1],
        reverse=True
    )

    results = []

    for index, score in ranked[:TOP_CHUNKS]:

        if score <= 0:
            continue

        item = documents[index].copy()

        item["score"] = float(score)

        results.append(item)

    return results


# =========================================================
# ASK MODEL
# =========================================================

def ask_llm(question, sources):

    if not sources:

        context = """
No relevant documents were found.

Do NOT invent information.
Tell the user that the available documents
do not contain enough information.
"""

    else:

        context_parts = []

        for number, source in enumerate(
            sources,
            start=1
        ):

            modified = datetime.fromtimestamp(
                source["modified"]
            ).strftime("%Y-%m-%d %H:%M")

            context_parts.append(
                f"""
SOURCE {number}
File: {source['file']}
Last modified: {modified}

CONTENT:
{source['text']}
"""
            )

        context = "\n".join(context_parts)

    system_prompt = """
You are a document-grounded local AI assistant.

Your job is to answer questions using ONLY
the provided document evidence.

IMPORTANT RULES:

1. Do not invent facts.
2. Do not rely on your general knowledge when
   the documents do not support an answer.
3. Compare the provided sources before answering.
4. If two sources disagree, explicitly say they
   conflict.
5. Prefer the newer document when the documents
   clearly refer to the same information and the
   newer document supersedes the old one.
6. If you cannot determine which source is newer
   or authoritative, say so.
7. If there is insufficient evidence, say:
   "I could not find enough information in the
   available documents."
8. Give a concise and direct answer.
"""

    user_prompt = f"""
DOCUMENT EVIDENCE:

{context}

USER QUESTION:

{question}

Now compare the evidence and answer the question.
"""

    payload = {
        "model": "local-model",

        "messages": [
            {
                "role": "system",
                "content": system_prompt
            },
            {
                "role": "user",
                "content": user_prompt
            }
        ],

        "temperature": 0.1,

        "max_tokens": 1000
    }

    response = requests.post(
        LLAMA_URL,
        json=payload,
        timeout=300
    )

    response.raise_for_status()

    data = response.json()

    return data["choices"][0]["message"]["content"]


# =========================================================
# REQUEST MODEL
# =========================================================

class Question(BaseModel):
    question: str


# Frontend routes removed. PHP will handle frontend serving.

# =========================================================
# ASK AI
# =========================================================

@app.post("/ask")
def ask_ai(request: Question):

    question = request.question.strip()

    if not question:
        raise HTTPException(
            status_code=400,
            detail="Question is empty."
        )

    sources = search_documents(
        question
    )

    try:

        answer = ask_llm(
            question,
            sources
        )

    except Exception as error:

        raise HTTPException(
            status_code=500,
            detail=f"AI server error: {error}"
        )

    return {
        "answer": answer,

        "sources": [
            {
                "file": source["file"],
                "score": round(
                    source["score"],
                    3
                )
            }

            for source in sources
        ]
    }


# =========================================================
# SUPABASE API ROUTES
# =========================================================

import os
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv(Path(__file__).parent / ".env")
url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_KEY")
supabase: Client = create_client(url, key)

class Announcement(BaseModel):
    title: str
    body: str
    category: str = 'news'
    audience: List[str] = ['students', 'faculty']
    status: str = 'published'

@app.post("/api/announcements")
def create_announcement(ann: Announcement):
    try:
        data = supabase.table("announcements").insert({
            "title": ann.title,
            "body": ann.body,
            "category": ann.category,
            "audience": ann.audience,
            "status": ann.status
        }).execute()
        return {"success": True, "data": data.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/announcements")
def get_announcements():
    try:
        data = supabase.table("announcements").select("*").order("created_at", desc=True).execute()
        return {"announcements": data.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/classes")
def get_classes(student_id: str):
    try:
        # Fetch the student's enrolled classes
        enrollments = supabase.table("student_classes").select("class_id").eq("student_id", student_id).execute()
        class_ids = [e['class_id'] for e in enrollments.data]
        
        if not class_ids:
            return {"classes": []}
            
        # Fetch the class details
        # Using in_ filter to get all classes matching the ids
        classes_data = supabase.table("classes").select("*").in_("id", class_ids).execute()
        return {"classes": classes_data.data}
    except Exception as e:
        # Fallback to mock data if the tables don't exist yet (user hasn't run the SQL)
        print(f"Error fetching classes (table might not exist): {e}")
        return {"classes": [
            {
                "id": "mock-1",
                "course_code": "CS101",
                "course_name": "Introduction to Programming",
                "instructor": "Prof. Smith",
                "room": "Room 301",
                "schedule_time": "09:00 AM - 10:30 AM",
                "schedule_days": "MWF"
            },
            {
                "id": "mock-2",
                "course_code": "MATH201",
                "course_name": "Advanced Calculus",
                "instructor": "Dr. Jones",
                "room": "Room 205",
                "schedule_time": "11:00 AM - 12:30 PM",
                "schedule_days": "TTh"
            }
        ]}

class Conversation(BaseModel):
    user_id: str
    title: str = "New Conversation"

@app.post("/api/conversations")
def create_conversation(conv: Conversation):
    try:
        data = supabase.table("chat_conversations").insert({
            "user_id": conv.user_id,
            "title": conv.title
        }).execute()
        return {"success": True, "data": data.data[0] if data.data else None}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/conversations")
def get_conversations(user_id: str):
    try:
        data = supabase.table("chat_conversations").select("*").eq("user_id", user_id).order("updated_at", desc=True).execute()
        return {"conversations": data.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ChatMessage(BaseModel):
    conversation_id: str
    role: str
    content: str
    sources: Optional[list] = None

@app.post("/api/messages")
def save_message(msg: ChatMessage):
    try:
        data = supabase.table("chat_messages").insert({
            "conversation_id": msg.conversation_id,
            "role": msg.role,
            "content": msg.content,
            "sources": msg.sources
        }).execute()
        
        # update conversation updated_at
        supabase.table("chat_conversations").update({"updated_at": "now()"}).eq("id", msg.conversation_id).execute()
        
        return {"success": True, "data": data.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/conversations/{conversation_id}/messages")
def get_messages(conversation_id: str):
    try:
        data = supabase.table("chat_messages").select("*").eq("conversation_id", conversation_id).order("created_at").execute()
        return {"messages": data.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class AttendanceCheckIn(BaseModel):
    student_id: str
    method: str = "qr_code"

@app.post("/api/attendance")
def check_in(data: AttendanceCheckIn):
    try:
        name = f"Student {data.student_id}"
        res = supabase.table("attendance_records").insert({
            "student_id": data.student_id,
            "student_name": name,
            "method": data.method
        }).execute()
        return {"success": True, "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class LoginRequest(BaseModel):
    email: str
    password: str

@app.post("/api/login")
def login_user(data: LoginRequest):
    try:
        if not data.email.endswith("@navotaspolytechniccollege.edu.ph"):
            return {"success": False, "message": "Only @navotaspolytechniccollege.edu.ph emails are allowed."}
            
        res = supabase.table("users").select("*").eq("email", data.email).execute()
        if not res.data:
            return {"success": False, "message": "Invalid email or password."}
            
        user = res.data[0]
        # In production this should be hashed, but as per requirements, we are checking plain text demo passwords
        if user["password_hash"] != data.password:
            return {"success": False, "message": "Invalid email or password."}
            
        return {
            "success": True, 
            "user": {
                "id": user["id"],
                "email": user["email"],
                "full_name": user["full_name"],
                "student_number": user["student_number"],
                "role": user["role"]
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =========================================================
# UPLOAD FILE
# =========================================================

@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...)
):

    allowed = [
        ".pdf",
        ".docx",
        ".txt",
        ".md",
        ".xlsx"
    ]

    extension = Path(
        file.filename
    ).suffix.lower()

    if extension not in allowed:

        raise HTTPException(
            status_code=400,
            detail="Unsupported file type."
        )

    filename = Path(
        file.filename
    ).name

    destination = DOCUMENTS_DIR / filename

    with destination.open("wb") as buffer:

        shutil.copyfileobj(
            file.file,
            buffer
        )

    # Immediately re-index documents
    rebuild_index()

    return {
        "success": True,
        "filename": filename,
        "message": "File uploaded and indexed."
    }


# =========================================================
# LIST DOCUMENTS
# =========================================================

@app.get("/documents")
def list_documents():

    files = []

    for path in DOCUMENTS_DIR.rglob("*"):

        if path.is_file():

            files.append({
                "name": path.name,
                "size": path.stat().st_size,
                "modified": datetime.fromtimestamp(
                    path.stat().st_mtime
                ).isoformat()
            })

    return {
        "documents": files
    }


# =========================================================
# STARTUP
# =========================================================

@app.on_event("startup")
def startup():

    rebuild_index()